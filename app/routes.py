from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for, flash, send_from_directory, send_file
from werkzeug.utils import secure_filename
from PIL import Image
import os
import uuid
import io
import zipfile
from datetime import datetime
from . import db
from .models import Admin, ClassGroup, Session, Flag
from .config import Config

main = Blueprint('main', __name__)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'txt', 'docx', 'mp4', 'webm', 'mov'}

@main.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(Config.UPLOAD_FOLDER, filename)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# --- Admin Routes ---
@main.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        password = request.form.get('password')
        admin = Admin.query.first()
        if admin and admin.check_password(password):
            session['admin_logged_in'] = True
            return redirect(url_for('main.admin_classroom'))
        flash('Invalid password')
    return render_template('admin_login.html')

@main.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('main.admin_login'))

@main.route('/admin/dashboard')
def admin_dashboard():
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    
    class_type = request.args.get('type', 'classmap')
    active_classes = ClassGroup.query.filter_by(is_active=True, class_type=class_type).all()
    past_classes = ClassGroup.query.filter_by(is_active=False, class_type=class_type).all()
    return render_template('admin_dashboard.html', 
                           active_classes=active_classes, 
                           past_classes=past_classes,
                           current_type=class_type)


@main.route('/admin/create_class', methods=['POST'])
def create_class():
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    name = request.form.get('name')
    class_type = request.form.get('class_type', 'classmap')
    
    if not name:
        name = f"Class {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    new_class = ClassGroup(name=name, class_type=class_type)
    db.session.add(new_class)
    db.session.commit()
    return redirect(url_for('main.admin_dashboard', type=class_type))

@main.route('/admin/close_class/<class_id>', methods=['POST'])
def close_class(class_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    c = ClassGroup.query.get_or_404(class_id)
    c.is_active = False
    db.session.commit()
    return redirect(url_for('main.admin_dashboard', type=c.class_type))

@main.route('/admin/class/<class_id>')
def admin_class(class_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    c = ClassGroup.query.get_or_404(class_id)
    active_sessions = Session.query.filter_by(class_id=class_id, is_active=True).all()
    past_sessions = Session.query.filter_by(class_id=class_id, is_active=False).all()
    return render_template('admin_class.html', class_group=c, active_sessions=active_sessions, past_sessions=past_sessions)

@main.route('/admin/class/<class_id>/create_session', methods=['POST'])
def create_session(class_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    c = ClassGroup.query.get_or_404(class_id)
    name = request.form.get('name')
    if not name:
        name = f"Session {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    new_session = Session(name=name, class_id=c.id)
    db.session.add(new_session)
    db.session.commit()
    return redirect(url_for('main.admin_class', class_id=c.id))

@main.route('/admin/close_session/<session_id>', methods=['POST'])
def close_session(session_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    s = Session.query.get_or_404(session_id)
    s.is_active = False
    db.session.commit()
    return redirect(url_for('main.admin_class', class_id=s.class_id))

@main.route('/admin/class/<class_id>/quiz_results')
def admin_class_quiz_results(class_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    c = ClassGroup.query.get_or_404(class_id)
    sessions = Session.query.filter_by(class_id=class_id).all()
    return render_template('class_quiz_results.html', class_group=c, sessions=sessions)

@main.route('/admin/settings')
def admin_settings():
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    return render_template('admin_settings.html')

@main.route('/admin/change_password', methods=['POST'])
def change_password():
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    new_password = request.form.get('new_password')
    if new_password:
        admin = Admin.query.first()
        admin.set_password(new_password)
        db.session.commit()
        flash('Password changed successfully.')
    return redirect(url_for('main.admin_settings'))

@main.route('/admin/reset_data', methods=['POST'])
def reset_data():
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    
    import shutil
    db.session.query(Flag).delete()
    db.session.query(Session).delete()
    db.session.query(ClassGroup).delete()
    db.session.commit()
    
    if os.path.exists(Config.UPLOAD_FOLDER):
        for filename in os.listdir(Config.UPLOAD_FOLDER):
            file_path = os.path.join(Config.UPLOAD_FOLDER, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print('Failed to delete %s. Reason: %s' % (file_path, e))
                
    flash('All data has been successfully reset.')
    return redirect(url_for('main.admin_settings'))

@main.route('/admin/export_markdown')
def export_markdown():
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    
    # Create a ZIP file in memory
    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Fetch all classes, sessions, and flags
        classes = ClassGroup.query.all()
        
        for c in classes:
            class_folder = secure_filename(c.name) or f"Class_{c.id[:8]}"
            
            for s in c.sessions:
                session_folder = secure_filename(s.name) or f"Session_{s.id[:8]}"
                
                for f in s.flags:
                    # Create markdown content
                    created_str = f.created_at.strftime('%Y-%m-%d %H:%M:%S')
                    safe_author = secure_filename(f.author_name) or "Participant"
                    filename = f"{f.id}_{safe_author}_{f.created_at.strftime('%Y%m%d_%H%M%S')}.md"
                    filepath = os.path.join(class_folder, session_folder, filename)
                    
                    md_content = f"""# Post by {f.author_name}
**Date:** {created_str}
**Class:** {c.name}
**Session:** {s.name}
"""
                    if f.x is not None and f.y is not None:
                        md_content += f"**Location:** ({f.x}, {f.y})\n"
                    
                    md_content += "\n---\n\n"
                    md_content += f.text_content if f.text_content else "*No text content*"
                    md_content += "\n\n---\n"
                    
                    if f.file_path:
                        md_content += f"**Attached File:** {f.file_path}\n"
                    
                    # Add to zip
                    zf.writestr(filepath, md_content)
        
        # If no flags found, add a placeholder
        if not Flag.query.first():
            zf.writestr("empty_export.txt", "No posts found to export.")

    memory_file.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        memory_file,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f"classroom_export_{timestamp}.zip"
    )

# --- Portal and Common Routes ---
@main.route('/')
def index():
    # Integrated portal for everyone
    return render_template('classroom_select.html', is_admin=session.get('admin_logged_in', False))

@main.route('/admin/classroom')
def admin_classroom():
    # Still keep this for explicit admin access
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    return render_template('classroom_select.html', is_admin=True)

# --- Participant Routes ---
@main.route('/classes')
def classes():
    # Participant dashboard showing filtered active classes
    class_type = request.args.get('type', 'classmap')
    active_classes = ClassGroup.query.filter_by(is_active=True, class_type=class_type).all()
    is_admin = session.get('admin_logged_in', False)
    return render_template('index.html', classes=active_classes, current_type=class_type, is_admin=is_admin)

@main.route('/class/<class_id>')
def view_class(class_id):
    c = ClassGroup.query.get_or_404(class_id)
    if not c.is_active and not session.get('admin_logged_in'):
        return "This class is closed.", 403
    active_sessions = Session.query.filter_by(class_id=class_id, is_active=True).all()
    is_admin = session.get('admin_logged_in', False)
    return render_template('class_sessions.html', class_group=c, sessions=active_sessions, is_admin=is_admin)

@main.route('/session/<session_id>')
def view_session(session_id):
    s = Session.query.get_or_404(session_id)
    if not s.is_active and not session.get('admin_logged_in'):
        return "This session is closed.", 403
        
    is_admin = session.get('admin_logged_in', False)
    if s.class_group.class_type == 'classquiz':
        return render_template('quiz_session.html', quiz_session=s, is_admin=is_admin)
    return render_template('map_session.html', map_session=s, is_admin=is_admin)

@main.route('/admin/session/<session_id>/export_quiz')
def export_quiz_excel(session_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    
    from openpyxl import Workbook
    from .models import QuizQuestion
    
    s = Session.query.get_or_404(session_id)
    questions = QuizQuestion.query.filter_by(session_id=session_id).order_by(QuizQuestion.index).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Quiz Questions"
    
    # Headers
    headers = ['Idx', 'Type', 'Question', 'Options (split by |)', 'Correct Answer']
    ws.append(headers)
    
    # Data
    for q in questions:
        ws.append([q.index, q.q_type, q.question, q.options, q.correct_answer])
    
    # Save to memory
    memory_file = io.BytesIO()
    wb.save(memory_file)
    memory_file.seek(0)
    
    filename = f"quiz_{secure_filename(s.name)}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        memory_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@main.route('/admin/session/<session_id>/import_quiz', methods=['POST'])
def import_quiz_excel(session_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    from openpyxl import load_workbook
    from .models import QuizQuestion
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        # Optionally clear existing questions? User didn't specify, but usually expected.
        # Let's keep existing for now or just append. 
        # Actually, let's clear existing to make it a "sync".
        QuizQuestion.query.filter_by(session_id=session_id).delete()
        
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[2]: continue # Skip if no question text
            
            new_q = QuizQuestion(
                session_id=session_id,
                index=row[0] if row[0] is not None else 0,
                q_type=row[1] if row[1] else 'choice',
                question=str(row[2]),
                options=str(row[3]) if row[3] else "",
                correct_answer=str(row[4]) if row[4] else ""
            )
            db.session.add(new_q)
        
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4().hex}_{filename}"
        filepath = os.path.join(Config.UPLOAD_FOLDER, unique_filename)
        
        # Ensure upload dir exists
        os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)
        file.save(filepath)
        
        # Create thumbnail if image
        thumbnail_path = None
        if filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'gif'}:
            try:
                img = Image.open(filepath)
                img.thumbnail((150, 150))
                thumb_filename = f"thumb_{unique_filename}"
                thumb_filepath = os.path.join(Config.UPLOAD_FOLDER, thumb_filename)
                img.save(thumb_filepath)
                thumbnail_path = f"uploads/{thumb_filename}"
            except Exception as e:
                print(f"Error creating thumbnail: {e}")
                
        return jsonify({
            'success': True, 
            'file_path': f"uploads/{unique_filename}",
            'thumbnail_path': thumbnail_path
        })
    return jsonify({'error': 'Invalid file type'}), 400
